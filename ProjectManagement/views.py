from datetime import datetime
from django.apps import apps
from django.http import HttpResponse, JsonResponse
from django.db import connection
import csv
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
from urllib.parse import quote

from django.db.models import Sum, Q
from django_filters.rest_framework import DjangoFilterBackend, FilterSet
from rest_framework import filters
from rest_framework.response import Response
from rest_framework import  status, permissions
from Core.Core.drf import generics
from django.contrib.auth import get_user_model

# from Permissions.models import Assignees
from ProjectManagement.models import PerformanceBankGuarantee
User = get_user_model()
from django_filters.rest_framework import DjangoFilterBackend, FilterSet
from django_filters import DateRangeFilter,DateFilter, TimeFilter, DateTimeFilter
from django_filters.filters import Filter
from django_filters import CharFilter
import django_filters
from Core.Core.permissions.permissions import GetPermission
from rest_framework.exceptions import NotFound
from django.shortcuts import get_object_or_404

from .serializers import *

from Masters.models import *



class ListFilter(Filter):
    def filter(self, qs, value):
        if not value:
            return qs

        self.lookup_expr = 'in'
        values = value.split(',')
        return super(ListFilter, self).filter(qs, values)
    

class ProjectFilter(FilterSet):
    date_range = DateRangeFilter(field_name='created_on')
    start_date = DateFilter(field_name='created_on',lookup_expr=('gte'),)
    end_date = DateFilter(field_name='created_on',lookup_expr=('lte'))
    company = django_filters.ModelChoiceFilter(
        queryset=apps.get_model('DynamicDjango', 'Company').objects.all()
    )
    class Meta:
        model = Project
        fields = ['start_date','end_date','date_range','status','tender','company','location','sourceportal',
                  'customer','order_placing_authority','manager','inspection_agency','is_performace_bank_guarantee',
                  'is_pre_dispatch_inspection','is_inspection_agency','current_authorized_status','authorized_status']



class ProjectItemFilter(FilterSet):
    date_range = DateRangeFilter(field_name='project__created_on')
    start_date = DateFilter(field_name='project__created_on',lookup_expr=('gte'),)
    end_date = DateFilter(field_name='project__created_on',lookup_expr=('lte'))

    company = django_filters.ModelChoiceFilter(
        field_name='project__company',
        queryset=apps.get_model('DynamicDjango', 'Company').objects.all()
    )

    location = django_filters.ModelChoiceFilter(
        field_name='project__location',
        queryset=apps.get_model('DynamicDjango', 'Location').objects.all()
    )

    customer = django_filters.ModelChoiceFilter(
        field_name='project__customer',
        queryset=Account.objects.filter(is_deleted=False)
    )

    tender = django_filters.ModelChoiceFilter(
        field_name='project__tender',
        queryset=Tender.objects.filter(is_deleted=False)
    )

    sourceportal = django_filters.ModelChoiceFilter(
        field_name='project__sourceportal',
        queryset=SourcePortal.objects.filter(is_deleted=False)
    )


    class Meta:
        model = ProjectItem
        fields = ['start_date','end_date','date_range','project', 'company', 'location', 'customer', 'tender', 'sourceportal', 'tenderitemmaster']
        


class ProjectCreate(generics.CreateAPIView):
    serializer_class = ProjectSerializer
    model = serializer_class.Meta.model
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')


class ProjectList(generics.ListAPIView):
    serializer_class = ProjectSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    model = serializer_class.Meta.model
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    search_fields = ['name', 'code']
    filterset_class = ProjectFilter
    ordering_fields = ['code', 'created_on']

    
    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(status=Project.PROJECT)

        return queryset.order_by('-created_on')
    
    
class ProjectMiniList(generics.ListAPIView):
    serializer_class = ProjectMiniSerializer
    model = serializer_class.Meta.model
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    filter_backends = [DjangoFilterBackend, filters.SearchFilter,]
    search_fields = ['name', ]
    filterset_class = ProjectFilter
    ordering_fields = ['code','created_on']
    screen_type = 'entry'
    
    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(status=Project.PROJECT)

        return queryset.order_by('-created_on')

   
    

class AssigneeProjectMiniList(generics.ListAPIView):
    serializer_class = ProjectMiniSerializer
    model = serializer_class.Meta.model
    queryset = model.objects.filter(is_deleted=False , status=Project.PROJECT ).order_by('-created_on')
    filter_backends = [DjangoFilterBackend, filters.SearchFilter,]
    search_fields = ['name', ]
    filterset_class = ProjectFilter
    ordering_fields = ['code','created_on']
    screen_type = 'entry'

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(status=Project.PROJECT)

        return queryset.order_by('-created_on')
    

    

class ProjectDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ProjectSerializer
    queryset = Project.objects.filter(is_deleted=False).order_by('-created_on')

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset

    def perform_destroy(self, instance):
        instance.is_deleted=True
        instance.save()


class ProjectUpdateView(generics.RetrieveUpdateAPIView):
    serializer_class = ProjectUpdateSerializer
    queryset = Project.objects.filter(is_deleted=False).order_by('-created_on')

    def get_queryset(self):
        user = self.request.user
        queryset = Project.objects.filter(is_deleted=False,)
        return queryset


class ProjectDocumentsCreate(generics.CreateAPIView):
    serializer_class = ProjectDocumentsSerializer
    model = serializer_class.Meta.model
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    

class ProjectDocumentsFilter(FilterSet):
    start_date = DateFilter(field_name='created_on',lookup_expr=('gte'),) 
    end_date = DateFilter(field_name='created_on',lookup_expr=('lte'))
    
    class Meta:
        model = ProjectDocuments
        fields = ['project', 'start_date', 'end_date',]


class ProjectDocumentsList(generics.ListAPIView):

    serializer_class = ProjectDocumentsSerializer
    model = serializer_class.Meta.model
    filter_backends = [DjangoFilterBackend, filters.SearchFilter,]
    filterset_class = ProjectDocumentsFilter
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    search_fields = ['project__name','code']
    ordering_fields = ['code','created_on']


class ProjectGroupFilter(FilterSet):
    date_range = DateRangeFilter(field_name='created_on')
    start_date = DateFilter(field_name='created_on',lookup_expr=('gte'),)
    end_date = DateFilter(field_name='created_on',lookup_expr=('lte'))

    class Meta:
        model = ProjectGroup
        fields = ['project','date_range','start_date', 'end_date' ]


class ProjectGroupCreate(generics.CreateAPIView):
    serializer_class = ProjectGroupSerializer
    model = serializer_class.Meta.model
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    

class ProjectGroupList(generics.ListAPIView):

    serializer_class = ProjectGroupSerializer
    model = serializer_class.Meta.model
    filter_backends = [DjangoFilterBackend, filters.SearchFilter,]
    filterset_class = ProjectGroupFilter
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    search_fields = ['name', 'project__name','code']
    ordering_fields = ['code','created_on']


class ProjectGroupMiniList(generics.ListAPIView):

    serializer_class = ProjectGroupMiniSerializer
    model = serializer_class.Meta.model
    filter_backends = [DjangoFilterBackend, filters.SearchFilter,]
    filterset_class = ProjectGroupFilter
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    search_fields = ['name', 'project__name','code']
    ordering_fields = ['code','created_on']



class ProjectGroupDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ProjectGroupSerializer
    model = serializer_class.Meta.model
    queryset = model.objects.filter(is_deleted = False)
        
    def perform_destroy(self, instance):
        instance.is_deleted=True
        instance.save()



class ProjectGroupUserFilter(FilterSet):

    class Meta:
        model = ProjectGroupUser
        fields = ['group', 'user',]


class ProjectGroupUserCreate(generics.CreateAPIView):
    # permission_classes = [permissions.IsAuthenticated]
    serializer_class = ProjectGroupUserSerializer
    model = serializer_class.Meta.model
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    

class ProjectGroupUserList(generics.ListAPIView):

    serializer_class = ProjectGroupUserSerializer
    model = serializer_class.Meta.model
    filter_backends = [DjangoFilterBackend, filters.SearchFilter,]
    filterset_class = ProjectGroupUserFilter
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    search_fields = ['group__name', 'group__project__name', 'user__username', 'code']
    ordering_fields = ['code','created_on']


class ProjectGroupUserMiniList(generics.ListAPIView):

    serializer_class = ProjectGroupUserMiniSerializer
    model = serializer_class.Meta.model
    filter_backends = [DjangoFilterBackend, filters.SearchFilter,]
    filterset_class = ProjectGroupUserFilter
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    search_fields = ['group__name', 'group__project__name', 'user__username', 'code']
    ordering_fields = ['code','created_on']


class ProjectGroupUserDetail(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ProjectGroupUserSerializer
    model = serializer_class.Meta.model
    queryset = model.objects.filter(is_deleted = False)

    def perform_destroy(self, instance):
        instance.is_deleted=True
        instance.save()



class ProjectDueDateDocumentFilter(FilterSet):
    start_date = DateFilter(field_name='created_on',lookup_expr=('gte'),) 
    end_date = DateFilter(field_name='created_on',lookup_expr=('lte'))
    
    class Meta:
        model = ProjectDueDateDocument
        fields = ['project', 'start_date', 'end_date',]


class ProjectDueDateDocumentCreate(generics.CreateAPIView):
    # permission_classes = [permissions.IsAuthenticated]
    serializer_class = ProjectDueDateDocumentSerializer
    model = serializer_class.Meta.model
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')


class ProjectDueDateDocumentList(generics.ListAPIView):

    serializer_class = ProjectDueDateDocumentSerializer
    model = serializer_class.Meta.model
    filter_backends = [DjangoFilterBackend, filters.SearchFilter,]
    filterset_class = ProjectDueDateDocumentFilter
    queryset = model.objects.filter(is_deleted=False ).order_by('-created_on')
    search_fields = ['project_name','code']
    ordering_fields = ['code','created_on']



class PerformanceBankGuaranteeFilter(FilterSet):
    date_range = DateRangeFilter(field_name='created_on')
    start_date = DateFilter(field_name='created_on',lookup_expr=('gte'),)
    end_date = DateFilter(field_name='created_on',lookup_expr=('lte'))


    class Meta:
        model = PerformanceBankGuarantee
        fields = ['date_range', 'start_date', 'end_date', 'project', 'authorized_status','current_authorized_status', ]


class PerformanceBankGuaranteeHistoryFilter(FilterSet):
    date_range = DateRangeFilter(field_name='created_on')
    start_date = DateFilter(field_name='created_on',lookup_expr=('gte'),)
    end_date = DateFilter(field_name='created_on',lookup_expr=('lte'))


    class Meta:
        model = PerformanceBankGuaranteeHistory
        fields = ['date_range', 'start_date', 'end_date','project','pbg' ]

class PerformanceBankGuaranteeList(generics.ListCreateAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = PerformanceBankGuaranteeSerializer
    queryset = PerformanceBankGuarantee.objects.filter(is_deleted=False).order_by('-created_on')
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = PerformanceBankGuaranteeFilter
    search_fields = ['number', 'value']
    ordering_fields = ['code', 'created_on']


class PerformanceBankGuaranteeDetail(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = PerformanceBankGuaranteeSerializer
    queryset = PerformanceBankGuarantee.objects.filter(is_deleted=False)

    def perform_destroy(self, instance):
        instance.is_deleted = True
        instance.save()
        

class PerformanceBankGuaranteeMiniList(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = PerformanceBankGuaranteeMiniSerializer
    queryset = PerformanceBankGuarantee.objects.filter(is_deleted=False).only('id', 'number', 'value').order_by('number')
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = PerformanceBankGuaranteeFilter
    search_fields = ['number', 'value']
    ordering_fields = ['code', 'created_on']


class PBGExtendedDuedateUpdate(generics.UpdateAPIView):
    permission_classes = [GetPermission('ProjectManagement.change_pbg_extended_due_date')]
    serializer_class = PBGExtendedUpdateSerializer
    queryset = PerformanceBankGuarantee.objects.filter(is_deleted=False,authorized_status=PerformanceBankGuarantee.APPROVED)


class PerformanceBankGuaranteeHistoryList(generics.ListCreateAPIView):
    serializer_class = PerformanceBankGuaranteeHistorySerializer
    queryset = PerformanceBankGuaranteeHistory.objects.filter(is_deleted=False).order_by('-created_on')
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = PerformanceBankGuaranteeHistoryFilter
    search_fields = ['number', 'value']
    ordering_fields = ['code', 'created_on']


class StockFilter(FilterSet):
    class Meta:
        model = Stock
        fields = ['project','warehouse','item','batch',]

        
class StockMovementFilter(FilterSet):
   
    class Meta:
        model = Stock
        fields = ['warehouse','item']


class StockView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = StockSerializer

    def get_queryset(self):
        # Extract parameters from the URL path and query parameters
        warehouse_id = self.kwargs.get('warehouse_id')
        item_id = self.kwargs.get('item_id')
        batch_id = self.kwargs.get('batch_id')
        project_id = self.kwargs.get('project_id', None)

        if project_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Project ID ", code=404)
        if warehouse_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Warehouse ID", code=404)
        if item_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Item ID", code=404)
        if batch_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Batch ID", code=404)

        # Start with the base queryset
        queryset = Stock.objects.all()
        print('---------------------queryset', queryset)

        # Apply filters based on URL parameters and query parameters
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        if item_id:
            queryset = queryset.filter(item_id=item_id)
        if batch_id:
            queryset = queryset.filter(batch_id=batch_id)

        print('---------------------filtered_queryset', queryset)
        return queryset

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        aggregated_data = queryset.aggregate(total_quantity=Sum('quantity'))

        total_quantity = aggregated_data['total_quantity'] or 0

        # Ensure total_quantity is not negative
        total_quantity = max(0, total_quantity)

        result = {
            "project": queryset.first().project.id if queryset.exists() else None,
            "warehouse": queryset.first().warehouse.id if queryset.exists() else None,
            "item": queryset.first().item.id if queryset.exists() else None,
            "batch": queryset.first().batch.id if queryset.exists() else None,
            "quantity": str(total_quantity)
        }

        # Return the result as a single object
        return Response(result, status=status.HTTP_200_OK)


class StockWithoutBatchView(generics.ListAPIView):

    permission_classes = [permissions.IsAuthenticated]
    serializer_class = StockSerializer


    def get_queryset(self):
        project_id = self.kwargs.get('project_id')
        warehouse_id = self.kwargs.get('warehouse_id')
        item_id = self.kwargs.get('item_id')

        print(f"project_id: {project_id}, warehouse_id: {warehouse_id}, item_id: {item_id}")

        if project_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Project ID ", code=404)
        if warehouse_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Warehouse ID", code=404)
        if item_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Item ID", code=404)

        # Start with the base queryset
        queryset = Stock.objects.all()
        print('---------------------queryset', queryset)

        # Apply filters based on URL parameters and query parameters
        if project_id:
            queryset = queryset.filter(project_id=project_id)
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
        if item_id:
            queryset = queryset.filter(item_id=item_id)
        # if batch_id:
        #     queryset = queryset.filter(batch_id=batch_id)

        print('---------------------filtered_queryset', queryset)
        return queryset
    

    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        aggregated_data = queryset.aggregate(total_quantity=Sum('quantity'))

        total_quantity = aggregated_data['total_quantity'] or 0
        # Ensure total_quantity is not negative
        total_quantity = max(0, total_quantity)

        result = {
            "project": queryset.first().project.id if queryset.exists() else None,
            "warehouse": queryset.first().warehouse.id if queryset.exists() else None,
            "item": queryset.first().item.id if queryset.exists() else None,
            # "batch": queryset.first().batch.id if queryset.exists() else None,
            "quantity": str(total_quantity)
        }

        # Return the result as a single object
        return Response(result, status=status.HTTP_200_OK)


class StockAgainstBatchView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = StockAgainstBatchSerializer

    def get_queryset(self):
        project_id = self.kwargs.get('project_id')
        warehouse_id = self.kwargs.get('warehouse_id')
        item_id = self.kwargs.get('item_id')

        print(f"Filtering for project_id={project_id}, warehouse_id={warehouse_id}, item_id={item_id}")

        if project_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Project ID ", code=404)
        if warehouse_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Warehouse ID", code=404)
        if item_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Item ID", code=404)

        queryset = Stock.objects.all()
        
        if project_id:
            queryset = queryset.filter(project_id=project_id)
            print(f"Filtered by project_id: {queryset.query}")
        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
            print(f"Filtered by warehouse_id: {queryset.query}")
        if item_id:
            queryset = queryset.filter(item_id=item_id)
            print(f"Filtered by item_id: {queryset.query}")

        grouped_queryset = queryset.values(
            'project_id', 
            'project__name', 
            'warehouse_id', 
            'warehouse__name',
            'batch_id', 
            'batch__name',
            'item_id', 
            'item__name', 
        ).annotate(total_quantity=Sum('quantity'))

        print(f"Final grouped queryset: {grouped_queryset.query}")
        return grouped_queryset

    def get(self, request, *args, **kwargs):
        print("Fetching queryset...")
        queryset = self.get_queryset()

        results = []
        for stock in queryset:
            total_quantity = stock['total_quantity'] or 0
            total_quantity = max(0, total_quantity)  # Ensure non-negative quantity

            print(f"Processing stock entry: {stock}")

            if total_quantity > 0:
                results.append({
                    "project": stock['project_id'],
                    "projectname": stock['project__name'],
                    "warehouse": stock['warehouse_id'],
                    "warehousename": stock['warehouse__name'],
                    "batch": stock['batch_id'],
                    "batchname": stock['batch__name'],
                    "item": stock['item_id'],
                    "itemname": stock['item__name'],
                    "quantity": str(total_quantity)
                })

        print(f"Total valid stock entries: {len(results)}")

        # Apply pagination
        paginator = self.pagination_class()
        paginated_queryset = paginator.paginate_queryset(results, request, view=self)

        total_count = len(results)
        print(f"Returning {len(paginated_queryset)} items out of {total_count} total.")

        return Response({
            'count': total_count,
            'results': paginated_queryset
        })
    

class StockItemView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ItemMiniSerializer

    def get_queryset(self):
        warehouse_id = self.kwargs.get('warehouse_id')
        project_id = self.kwargs.get('project_id', None)

        if project_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Project ID ", code=404)
        if warehouse_id == 'undefined':
            raise NotFound(detail="Invalid format provided for Warehouse ID", code=404)

        queryset = Stock.objects.all()

        if warehouse_id:
            queryset = queryset.filter(warehouse_id=warehouse_id)
        if project_id:
            queryset = queryset.filter(project_id=project_id)

        return queryset


    def get(self, request, *args, **kwargs):
        queryset = self.get_queryset()

        item_quantities = (
            queryset.values('item_id')
            .annotate(total_quantity=Sum('quantity'))
            .filter(total_quantity__gt=0)  # Only keep items with quantity > 0
        )

        # Fetch distinct item objects from the aggregated item IDs
        item_ids = [entry['item_id'] for entry in item_quantities]
        items = Item.objects.filter(id__in=item_ids)

        # Serialize the distinct items
        serializer = self.get_serializer(items, many=True)
        results = serializer.data

        # Apply pagination
        paginator = self.pagination_class()
        paginated_queryset = paginator.paginate_queryset(results, request, view=self)

        total_count = len(results)
        print(f"Returning {len(paginated_queryset)} items out of {total_count} total.")

        return Response({
            'count': total_count,
            'results': paginated_queryset
        }, status=status.HTTP_200_OK)

    
class ProjectItemsByProjectView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['tenderitemmaster__name', 'tenderitemmaster__code']
    ordering_fields = ['tenderitemmaster__code', 'created_on']
    pagination_class = None

    def get(self, request, *args, **kwargs):
        project_id = self.kwargs.get('project_id')
        print(f'ProjectItemsByProjectView called with project_id: {project_id}')
        if not project_id:
            return Response([], status=status.HTTP_200_OK)
        
        # Get project items with quantities
        project_items = ProjectItem.objects.filter(
            project_id=project_id, 
            is_deleted=False
        ).select_related('tenderitemmaster').filter(
            tenderitemmaster__is_deleted=False,
            tenderitemmaster__product_type=Item.PRODUCT
        ).exclude(tenderitemmaster__authorized_status=3)
        
        print(f'Found {project_items.count()} project items for project {project_id}')
        
        # Apply search filter if provided
        search = request.GET.get('search')
        if search:
            project_items = project_items.filter(
                Q(tenderitemmaster__name__icontains=search) |
                Q(tenderitemmaster__code__icontains=search)
            )
        
        # Transform to include quantity
        results = []
        for project_item in project_items:
            item_data = {
                'id': project_item.tenderitemmaster.id,
                'name': project_item.tenderitemmaster.name,
                'code': project_item.tenderitemmaster.code,
                'mrp': project_item.tenderitemmaster.mrp,
                'rate': project_item.tenderitemmaster.rate,
                'quantity': float(project_item.quantity) if project_item.quantity else 0
            }
            print(f'Adding item: {item_data}')
            results.append(item_data)
        
        print(f'Returning {len(results)} items')
        return Response(results, status=status.HTTP_200_OK)


class StockLedgerView(generics.ListAPIView):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = StockSerializer
    filterset_class = StockFilter
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    
    def get(self, request, *args, **kwargs):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")
        warehouse_id = request.GET.get('warehouse_id')
        project_id = request.GET.get('project_id')
        item_id = request.GET.get('item_id')
        page_size = int(request.GET.get('page_size', 10))  # Default 10
        page_no = int(request.GET.get('page', 1))  # Default 1
        view_type = int(request.GET.get('view_type', 1))  # Default 1

        # Calculate the OFFSET
        offset = (page_no - 1) * page_size
        
        filter_q = ''
        if warehouse_id:
            filter_q += f""" AND "stock_view".warehouse_id='{warehouse_id}'"""
        if project_id:
            filter_q += f""" AND "stock_view".project_id='{project_id}'"""
        if item_id:
            filter_q += f""" AND "stock_view".item_id='{item_id}'"""

        try:
            if start_date:
                start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            else:
                start_date = datetime.date.today()

            if end_date:
                end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
            else:
                end_date = datetime.date.today()
        except Exception as e:
            print(f"Error parsing date: {e}")
            start_date = datetime.date.today()
            end_date = datetime.date.today()

        start_date_str = start_date.strftime("%Y-%m-%d")
        end_date_str = end_date.strftime("%Y-%m-%d")

        print("start_date:", start_date_str)
        print("end_date:", end_date_str)

        query = """
            WITH stock_data AS (
                SELECT
                    2 AS "os_seq",
                    "stock_view"."doc_code" AS "os_doc_code",
                    "stock_view"."date" AS "os_date",
                    "stock_view"."item_id" AS "os_item_id",
                    "stock_view"."warehouse_id" AS "os_warehouse_id",
                    "stock_view"."project_id" AS "os_project_id",
                    SUM("stock_view"."quantity") FILTER (WHERE ("stock_view"."quantity" > 0)) AS "os_received_quantity",
                    SUM("stock_view"."quantity") FILTER (WHERE ("stock_view"."quantity" < 0)) AS "os_issued_quantity",
                    SUM("stock_view"."quantity") AS "row_balance"
                FROM "stock_view"
                WHERE "stock_view"."date" >= '{start_date}' and "stock_view"."date" <= '{end_date}' {filter_q}
                GROUP BY "stock_view"."doc_code", "stock_view"."date",
                        "stock_view"."project_id", "stock_view"."warehouse_id", "stock_view"."item_id"
                       
            ),
            all_items AS (
                -- Get all unique item IDs from transaction data
                SELECT DISTINCT
                    "os_item_id",
                    "os_warehouse_id",
                    "os_project_id"
                FROM stock_data
            ),
            opening_stock AS (
                -- Opening stock for all items (Using IDs instead of Names)
                SELECT
                    1 AS "os_seq",
                    'Opening Stock' AS "os_doc_code",
                    NULL AS "os_date",
                    ai."os_item_id",
                    ai."os_warehouse_id",
                    ai."os_project_id",
                    COALESCE(
                        (SELECT SUM("stock_view"."quantity")
                        FROM "stock_view"
                        WHERE "stock_view"."date" < '{start_date}' 
                        AND "stock_view"."item_id" = ai."os_item_id"
                        AND "stock_view"."warehouse_id" = ai."os_warehouse_id"
                        AND "stock_view"."project_id" = ai."os_project_id"),
                        0
                    ) AS "os_received_quantity",
                    0 AS "os_issued_quantity",
                    COALESCE(
                        (SELECT SUM("stock_view"."quantity")
                        FROM "stock_view"
                        WHERE "stock_view"."date" < '{start_date}'
                        AND "stock_view"."item_id" = ai."os_item_id"
                        AND "stock_view"."warehouse_id" = ai."os_warehouse_id"
                        AND "stock_view"."project_id" = ai."os_project_id"),
                        0
                    ) AS "row_balance"
                FROM all_items ai
                
                UNION ALL
                
                SELECT * FROM stock_data
            )
            SELECT 
                os_seq,
                os_doc_code,
                os_date,
                i.name AS os_item_name,
                w.name AS os_warehouse_name,
                p.name AS os_project_name,
                os_received_quantity,
                os_issued_quantity,
                SUM(row_balance) OVER (
                    PARTITION BY os_warehouse_id, os_project_id, os_item_id 
                    ORDER BY os_seq, os_date NULLS FIRST
                    ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                ) AS os_balance_quantity
            FROM opening_stock
            LEFT JOIN public."Masters_item" i ON opening_stock.os_item_id = i.id
            LEFT JOIN public."DynamicDjango_warehouse" w ON opening_stock.os_warehouse_id = w.id
            LEFT JOIN public."ProjectManagement_project" p ON opening_stock.os_project_id = p.id
            ORDER BY os_warehouse_name,  os_item_name, os_seq

        """

        formatted_query = query.format(start_date= start_date, end_date= end_date, filter_q= filter_q)
        keys = [
            "Id", "Code", "Date","Item Name",
            "Warehouse Name", "Location Name","Received Quantity", 
            "Issued Quantity", "Balance Quantity","Project Name"
        ]
        
        if view_type == 2:
            results = []

            with connection.cursor() as cursor:
                cursor.execute(formatted_query)
                results = cursor.fetchall()
                
            # json_output = [dict(zip(keys, row)) for row in results] 
                # Define the CSV response
                
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="stock_data.csv"'

            # Create CSV writer
            writer = csv.writer(response)

            # Write CSV Header
            writer.writerow(keys)

            # Write CSV Data
            for row in results:
                writer.writerow(row)

            return response
        
        elif view_type == 3:  # Simple Excel Export
            # Fetch all results
            results = []
            with connection.cursor() as cursor:
                cursor.execute(formatted_query)
                results = cursor.fetchall()
            
            # Create a workbook and select the active worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stock Ledger"
            
            # Write header row
            header_style = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            # Header row at row 1
            for col_idx, header in enumerate(keys, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_style
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format column widths
            for col_idx, header in enumerate(keys, 1):
                column_width = max(len(str(header)), 15)  # Minimum width of 15
                ws.column_dimensions[get_column_letter(col_idx)].width = column_width
            
            # Alternating row colors
            even_row_fill = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
            received_qty_idx = keys.index("Received Quantity") + 1 if "Received Quantity" in keys else 8  # Default to column 8
            #received_bags_idx = keys.index("Received No of Bags") + 1 if "Received No of Bags" in keys else 9
            # Find index positions for key columns
            item_name_idx = keys.index("Item Name") + 1 if "Item Name" in keys else 5  # Default to column 5
            balance_qty_idx = keys.index("Balance Quantity") + 1 if "Balance Quantity" in keys else 12  # Default to column 12
            #balance_bags_idx = keys.index("Balance No of Bags") + 1 if "Balance No of Bags" in keys else 13  # Default to column 13
            
            # Process result rows
            opening_stock_items = {}  # Track opening stock balances by item
            
            # First pass - collect all balance quantities for opening stock items
            for row in results:
                item_name = row[item_name_idx - 1]  # Adjust for 0-based index in row data
                is_opening_stock = (row[0] == 1)
                
                if is_opening_stock:
                    # Note: balance quantity and bags are at fixed indices in the row data
                    # Getting the balance quantity and bags values
                    balance_qty = row[balance_qty_idx - 1] if len(row) >= balance_qty_idx else 0
                    #balance_bags = row[balance_bags_idx - 1] if len(row) >= balance_bags_idx else 0
                    opening_stock_items[item_name] = (balance_qty,)
            
            # Second pass - generate the Excel sheet
            for row_idx, row in enumerate(results, 2):  # Start from row 4 (after title and header)
                # Apply alternating row colors
                row_fill = even_row_fill if row_idx % 2 == 0 else None
                
                # Check if this is an opening stock row
                is_opening_stock = (row[0] == 1)
                item_name = row[item_name_idx - 1]  # Adjust for 0-based index in row data
                
                # Write data row with timezone fix
                for col_idx, value in enumerate(row, 1):
                    # Fix timezone issue - remove timezone info from datetime objects
                    if isinstance(value, datetime.datetime) and value.tzinfo is not None:
                        value = value.replace(tzinfo=None)
                    
                    # For opening stock rows, handle specially
                    if is_opening_stock:
                        if col_idx == item_name_idx:
                            # Item name column - keep original value
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        elif col_idx == balance_qty_idx:
                            # Balance quantity - use the stored value
                            balance_qty = opening_stock_items.get(item_name, (0))[0]
                            cell = ws.cell(row=row_idx, column=col_idx, value=balance_qty)
                        # elif col_idx == balance_bags_idx:
                        #     # Balance bags - use the stored value
                        #     balance_bags = opening_stock_items.get(item_name, (0, 0))[1]
                        #     cell = ws.cell(row=row_idx, column=col_idx, value=balance_bags)
                        elif col_idx == received_qty_idx:
                            # Also populate Received Quantity for opening stock
                            balance_qty = opening_stock_items.get(item_name, (0))[0]
                            cell = ws.cell(row=row_idx, column=col_idx, value=balance_qty)
                        # elif col_idx == received_bags_idx:
                        #     # Also populate Received No of Bags for opening stock
                        #     balance_bags = opening_stock_items.get(item_name, (0, 0))[1]
                        #     cell = ws.cell(row=row_idx, column=col_idx, value=balance_bags)
                        elif col_idx == 2:  # Code column
                            cell = ws.cell(row=row_idx, column=col_idx, value="Opening Stock")
                        else:
                            # Other columns - leave empty for merging
                            cell = ws.cell(row=row_idx, column=col_idx, value="")
                    else:
                        # Regular row - process normally
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Apply formatting
                    cell.border = border
                    
                    if col_idx in [8, 9, 10, 11, 12, 13]:
                        cell.number_format = '#,##0.00'
                    
                    # Apply date formatting for date columns
                    if col_idx == 3 and value:  # Date column
                        cell.number_format = 'dd-mm-yyyy'
                    
                    # Apply fill for alternating rows
                    if row_fill:
                        cell.fill = row_fill
                
                # Special formatting for opening stock rows and merge cells
                if is_opening_stock:
                    # Apply bold font to opening stock rows
                    for col_idx in range(1, len(keys) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Define merge regions for opening stock rows
                    # Merge from column 2 to item_name_idx-1
                    if item_name_idx > 2:
                        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=item_name_idx-1)
                       
                    if received_qty_idx - item_name_idx > 1:
                        ws.merge_cells(start_row=row_idx, start_column=item_name_idx+1, end_row=row_idx, end_column=received_qty_idx-1)
                   
                    if "Issued Quantity" in keys:
                        issued_qty_idx = keys.index("Issued Quantity") + 1
                        if issued_qty_idx > 1:
                            ws.merge_cells(start_row=row_idx, end_row=row_idx, end_column=issued_qty_idx-1)#start_column=received_bags_idx+1
           
            ws.freeze_panes = 'A2'
            
            # Save Excel file to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            file_name = 'StockLedger.xls'
            file_path = os.path.join(settings.MEDIA_ROOT, file_name)
            wb.save(file_path)
    
            return JsonResponse({"file_name": file_name})
            
            
        
        paginated_query = f"""
            SELECT * FROM (
                {formatted_query}
            ) AS subquery
            LIMIT {page_size} OFFSET {offset}
        """
        results = None

        with connection.cursor() as cursor:
            cursor.execute(paginated_query)
            results = cursor.fetchall()

        count_query="""
            select count(*) from ({query}) as subquery
        """
        count_results = None
        count_query = count_query.format(query = formatted_query)

        with connection.cursor() as cursor:
            cursor.execute(count_query)
            count_results = cursor.fetchone()[0]  # Extract the count from the tuple
            

        # Convert list to JSON objects
        json_output = [dict(zip(keys, row)) for row in results] 
        
        response_data = {
            'results': json_output,
            'count': count_results
        }
        return Response(response_data, status=status.HTTP_200_OK)
 